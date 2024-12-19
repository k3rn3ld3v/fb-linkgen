import tkinter as tk
from tkinter import ttk, messagebox
import re
import webbrowser
import pandas as pd
from datetime import datetime
from urllib.parse import urlparse, parse_qs
import requests
from bs4 import BeautifulSoup
import json
from openpyxl.styles import Border, Side, Alignment, Font
import random
import string

class FacebookLinkGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Facebook Link Generator")
        self.root.geometry("800x600")
        
        # Style
        self.root.configure(bg='#f0f2f5')
        self.style = ttk.Style()
        self.style.configure('TButton', padding=5)
        self.style.configure('TEntry', padding=5)
        self.style.configure('Export.TButton', 
                           background='#1877f2',
                           foreground='white',
                           font=('Helvetica', 10, 'bold'),
                           padding=10)
        
        # Main container
        main_container = ttk.Frame(root, padding="20")
        main_container.pack(fill='both', expand=True)
        
        # Input Frame
        input_frame = ttk.LabelFrame(main_container, text="Profile Information", padding="10")
        input_frame.pack(fill='x', pady=(0, 20))
        
        # URL Entry
        ttk.Label(input_frame, text="Enter Facebook Profile URL:").pack(fill='x')
        self.url_entry = ttk.Entry(input_frame, width=50)
        self.url_entry.pack(fill='x', pady=5)
        
        # Extract button
        self.extract_button = ttk.Button(input_frame, text="Extract Information", 
                                         command=self.extract_user_info,
                                         style='TButton')
        self.extract_button.pack(pady=5)
        
        # Info display frame
        info_frame = ttk.Frame(input_frame)
        info_frame.pack(fill='x', pady=5)
        
        # User info display
        ttk.Label(info_frame, text="Username: ").grid(row=0, column=0, sticky='e')
        self.username_entry = ttk.Entry(info_frame, textvariable=tk.StringVar(), state='readonly')
        self.username_entry.grid(row=0, column=1, sticky='w')
        
        ttk.Label(info_frame, text="User ID: ").grid(row=1, column=0, sticky='e')
        self.user_id_entry = ttk.Entry(info_frame, textvariable=tk.StringVar(), state='readonly')
        self.user_id_entry.grid(row=1, column=1, sticky='w')
        
        # Buttons Frame with Grid
        self.buttons_frame = ttk.LabelFrame(main_container, text="Facebook Links", padding="10")
        self.buttons_frame.pack(fill='both', expand=True)
        
        # Configure grid columns
        for i in range(3):
            self.buttons_frame.columnconfigure(i, weight=1)
        
        # Define endpoints
        self.endpoints = {
            "Photos Of": "https://www.facebook.com/{}/photos_of",
            "Videos Of": "https://www.facebook.com/{}/videos_of",
            "Stories Of": "https://www.facebook.com/stories/{}",
            "Groups": "https://www.facebook.com/{}/groups",
            "Events Joined": "https://www.facebook.com/{}/events",
            "Games": "https://www.facebook.com/{}/games",
            "Apps": "https://www.facebook.com/{}/apps",
            "Liked Photos": "https://www.facebook.com/{}/photos_liked",
            "Liked Videos": "https://www.facebook.com/{}/videos_liked",
            "Places Visited": "https://www.facebook.com/{}/places"
        }
        
        self.create_endpoint_buttons()
        
        # Export Button Frame
        export_frame = ttk.Frame(main_container)
        export_frame.pack(fill='x', pady=20)
        
        # Custom styled Export Button
        self.export_button = tk.Button(
            export_frame,
            text="Export All Links to Excel",
            command=self.export_to_excel,
            bg='#1877f2',
            fg='white',
            font=('Helvetica', 10, 'bold'),
            relief='raised',
            padx=20,
            pady=10,
            state='disabled'  # Initially disabled
        )
        self.export_button.pack(side='right')

    def extract_user_info(self):
        url = self.url_entry.get().strip()
        if not url:
            messagebox.showerror("Error", "Please enter a Facebook profile URL")
            return
        
        try:
            # Disable the extract button
            self.extract_button.config(state='disabled')
            self.root.update_idletasks()
            
            # Clean and standardize the URL
            if not url.startswith(('http://', 'https://')):
                url = 'https://' + url
            
            # Handle both www.facebook.com and facebook.com
            if 'facebook.com' not in url:
                messagebox.showerror("Error", "Please enter a valid Facebook URL")
                self.extract_button.config(state='normal')
                return None
                
            # Extract the user ID using BeautifulSoup
            user_id = self.get_facebook_user_id(url)
            if user_id == "Unknown":
                messagebox.showerror("Error", "User not found or given URL is not correct")
                self.extract_button.config(state='normal')
                return None
            
            # Extract the path after facebook.com
            parsed = urlparse(url)
            path = parsed.path.strip('/')
            parts = path.split('/')
            
            # Remove common URL parts
            parts = [p for p in parts if p and p not in ['profile.php', 'people']]
            
            if not parts:
                messagebox.showerror("Error", "Could not extract profile information from URL")
                self.extract_button.config(state='normal')
                return None
            
            # Handle username in URL
            username = parts[-1] if parts else ""
            
            self.username_entry.config(state='normal')
            self.username_entry.delete(0, tk.END)
            self.username_entry.insert(0, username)
            self.username_entry.config(state='readonly')
            
            self.user_id_entry.config(state='normal')
            self.user_id_entry.delete(0, tk.END)
            self.user_id_entry.insert(0, user_id)
            self.user_id_entry.config(state='readonly')
            
            # Re-enable the extract button
            self.extract_button.config(state='normal')
            # Enable the export button since data is fetched
            self.export_button.config(state='normal')
            return user_id
            
        except Exception as e:
            self.extract_button.config(state='normal')
            messagebox.showerror("Error", f"Error processing URL: {str(e)}")
            return None

    def get_facebook_user_id(self, url):
        html = requests.get(url).text
        soup = BeautifulSoup(html, 'html.parser')
        meta = soup.find('meta', {'property': 'al:ios:url'})
        if meta:
            content = meta['content']
            # Extract the numeric ID after 'fb://profile/'
            return content.split('/')[-1]
        return "Unknown"

    def create_endpoint_buttons(self):
        row = 0
        col = 0
        for name, url_template in self.endpoints.items():
            btn = ttk.Button(
                self.buttons_frame,
                text=name,
                command=lambda url_temp=url_template: self.open_url(url_temp)
            )
            btn.grid(row=row, column=col, pady=5, padx=5, sticky='nsew')
            
            col += 1
            if col > 2:  # 3 columns
                col = 0
                row += 1

    def open_url(self, url_template):
        user_id = self.user_id_entry.get()
        if not user_id:
            messagebox.showerror("Error", "Please extract user information first")
            return
        
        url = url_template.format(user_id)
        webbrowser.open(url)

    def export_to_excel(self):
        user_id = self.user_id_entry.get()
        if not user_id:
            messagebox.showerror("Error", "Please extract user information first")
            return
        
        username = self.username_entry.get()
        given_url = self.url_entry.get().strip()
        date_str = datetime.now().strftime('%d-%b-%Y')
        random_suffix = ''.join(random.choices(string.ascii_letters + string.digits, k=4))
        filename = f'fb-linkgen_{username}_{date_str}_{random_suffix}.xlsx'
        
        data = {
            'Link Type': [],
            'URL': []
        }
        
        for name, url_template in self.endpoints.items():
            data['Link Type'].append(name)
            data['URL'].append(url_template.format(user_id))
        
        df = pd.DataFrame(data)
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Links', startrow=5, index=False)
                workbook = writer.book
                worksheet = writer.sheets['Links']
                
                # Turn off gridlines
                worksheet.sheet_view.showGridLines = False
                
                # Set zoom level to 150%
                worksheet.sheet_view.zoomScale = 150
                
                # Write the given URL, username, and user ID
                worksheet['A1'] = 'Given URL:'
                worksheet['B1'] = given_url
                worksheet['A2'] = 'Username:'
                worksheet['B2'] = username
                worksheet['A3'] = 'UserID:'
                worksheet['B3'] = user_id
                
                # Apply styles
                for cell in ['A1', 'A2', 'A3']:
                    worksheet[cell].alignment = Alignment(horizontal='right')
                    worksheet[cell].font = Font(bold=True)
                for cell in ['B1', 'B2', 'B3']:
                    worksheet[cell].alignment = Alignment(horizontal='left')
                
                # Apply borderlines
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                for row in worksheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2):
                    for cell in row:
                        cell.border = thin_border
                for row in worksheet.iter_rows(min_row=5, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border
                
                # Autofit columns
                for column_cells in worksheet.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2
                
            messagebox.showinfo("Success", f"Links exported to {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Error exporting to Excel: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = FacebookLinkGenerator(root)
    root.mainloop()
